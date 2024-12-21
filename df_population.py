import pandas as pd
import os
from sqlalchemy import create_engine
from openpyxl import load_workbook
from dotenv import load_dotenv

load_dotenv()

# Configuration de la base de données
db_user = os.getenv("DB_USER")
db_password = os.getenv("DB_PASSWORD")
db_host = os.getenv("DB_HOST")
db_port = os.getenv("DB_PORT")
db_name = os.getenv("DB_NAME")

# Création de l'engine SQLAlchemy
db_url = f"postgresql+psycopg2://{db_user}:{db_password}@{db_host}:{db_port}/{db_name}"
engine = create_engine(db_url)

# Test de la connexion à la base de données
try:
    with engine.connect() as connection:
        print("Connexion réussie à la base de données !")
except Exception as e:
    print(f"Erreur de connexion: {e}")

# Fonction pour charger les données d'un fichier Excel
def charger_donnees_population(path_fichier, colonne_cible, nom_ville, plage_cells='A8:C12'):
    dossier = './DataLake'
    filepath = os.path.join(dossier, path_fichier)
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"Fichier introuvable : {filepath}")
    wb = load_workbook(filepath, data_only=True)
    ws = wb['Comparateur de territoires']
    data = [[cell.value for cell in row] for row in ws[plage_cells]]
    df = pd.DataFrame(data)
    # Définir les colonnes avec la première ligne
    df.columns = df.iloc[0]
    df = df.iloc[1:].reset_index(drop=True)
    df = df.set_index('Indicateurs')
    df_cible = df[colonne_cible].to_frame().T
    df_cible = df_cible.reset_index(drop=True)
    df_cible['Nom ville'] = nom_ville  # Ajouter la colonne manquante
    return df_cible

try:
    # Charger les données pour Argenteuil
    df_pop_argenteuil = charger_donnees_population(
        path_fichier='./Données démographiques/data_argenteuil_demog.xlsx',
        colonne_cible='Argenteuil',
        nom_ville='Argenteuil'
    )

    # Charger les données pour Versailles
    df_pop_versailles = charger_donnees_population(
        path_fichier='./Données démographiques/data_versailles_demog.xlsx',
        colonne_cible='Versailles',
        nom_ville='Versailles'
    )

    # Concaténer les deux DataFrames
    df_pop_comparaison = pd.concat([df_pop_argenteuil, df_pop_versailles])

    # Sélectionner les colonnes d'intérêt
    colonnes_interet = ['Nom ville', 'Population', 'Densité de population (hab/km²)', 'Superficie (km²)', 'Nombre de ménages']
    df_pop_comparaison = df_pop_comparaison[colonnes_interet]

    # Renommer les noms de colonnes
    df_pop_comparaison = df_pop_comparaison.rename(
        columns={
            'Nom ville': 'nom_ville',
            'Population':'nb_population',
            'Densité de population (hab/km²)': 'densite_de_pop',
            'Superficie (km²)': 'superficie',
            'Nombre de ménages': 'nombre_de_menage'
        }
    )

    # Vérifier les données avant insertion
    print(df_pop_comparaison.head())

    # Insérer dans la base de données
    df_pop_comparaison.to_sql('dim_population', engine, if_exists='replace', index=False)
    print("Données insérées avec succès dans la table dim_population.")
except Exception as e:
    print(f"Erreur : {e}")
